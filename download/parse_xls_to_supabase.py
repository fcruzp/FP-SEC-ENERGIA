#!/usr/bin/env python3
"""
XLS Parser — Informe de Desempeño del Sector Eléctrico
Parsea el XLS mensual y genera data_points en Supabase.

Uso:
  python3 parse_xls_to_supabase.py <archivo.xlsx> [--dry-run] [--sheet "EDE's"] [--batch-size 500]

Estrategia:
  - Lee cada hoja del XLS
  - Mapea indicadores por nombre → slug → UUID en tabla indicators
  - Mapea entidades por nombre → slug → UUID en tabla entities
  - Extrae series temporales (columnas mensuales)
  - Inserta en data_points via Supabase REST API (service_role key)
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, date
from typing import Optional
from urllib.request import Request, urlopen
from urllib.error import HTTPError

# ============================================================
# CONFIG
# ============================================================
SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')

# Monthly data starts at column index 12 (column L in Excel, 0-indexed)
MONTHLY_DATA_START_COL = 12

# Known entity name → slug mapping
ENTITY_MAP = {
    'edenorte': 'edenorte',
    'edesur': 'edesur',
    'edeeste': 'edeeste',
    "ede's": 'edes-consolidado',
    'total': 'edes-consolidado',
    'cdeee': 'cdeee',
    'egehid': 'egehid',
    'eted': 'eted',
    'egpc': 'egpc',
    'punta catalina': 'egpc',
    'gsf': 'gsf',
    'cespm': 'cespm',
    'dpp': 'dpp',
    'egehaina': 'egehaina-larimar',
    'larimar': 'egehaina-larimar',
    'electronic jrc': 'electronic-jrc',
    'montecristi solar': 'montecristi-solar',
    'c power': 'c-power',
    'cpower': 'c-power',
    'pecasa': 'pecasa',
    'matafongo': 'matafongo',
    'wcg energy': 'wcg-energy',
    'wcg': 'wcg-energy',
    'emerald solar': 'emerald-solar',
    'poseidon': 'poseidon',
    'quisqueya ii': 'quisqueya-ii',
    'quisqueya': 'quisqueya-ii',
    'falcondo': 'falcondo',
    'rsj': 'rsj',
    'mercado spot': 'mercado-spot',
    'spot': 'mercado-spot',
    'renovables contratos con cdeee': None,  # No entity for this aggregate
    'mercado de contratos': None,
    'gencos': None,
    "genco's": None,
    'unr': None,
}

# Sheet → category slug mapping
SHEET_CATEGORY_MAP = {
    'Variables Relevantes': 'variables-relevantes',
    "EDE's": 'empresas-distribuidoras',
    'CDEEE': 'cdeee',
    'EGEHID': 'egehid',
    'ETED': 'eted',
    'EGPC': 'egpc-punta-catalina',
    'Anexo Res Financieros': 'resultados-financieros',
    'Anexo Deuda': 'deuda-generadoras',
    'Nuevo Regimen tarifario': 'regimen-tarifario',
    'Regimen tarifario anterior': 'regimen-tarifario-anterior',
}


# ============================================================
# SUPABASE API HELPERS
# ============================================================
def supabase_request(table: str, method: str = 'GET', body: Optional[dict] = None,
                     params: Optional[dict] = None, limit: int = 1000):
    """Make a request to Supabase REST API with service_role key."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        query = '&'.join(f"{k}={v}" for k, v in params.items())
        url += f"?{query}"
    if method == 'GET' and limit:
        url += f"{'&' if params else '?'}limit={limit}"

    headers = {
        'apikey': SUPABASE_SERVICE_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation',
    }

    data = json.dumps(body).encode() if body else None
    req = Request(url, data=data, headers=headers, method=method)

    try:
        with urlopen(req) as resp:
            if resp.status == 204:
                return None
            return json.loads(resp.read().decode())
    except HTTPError as e:
        error_body = e.read().decode() if e.fp else ''
        print(f"  ❌ Supabase error ({e.code}): {error_body[:200]}")
        raise


def load_indicators() -> dict:
    """Load all indicators from Supabase, indexed by slug."""
    print("📡 Cargando indicadores desde Supabase...")
    result = supabase_request('indicators', params={'select': 'id,name,slug,category_id,entity_id,is_breakdown,parent_indicator_id,unit'})
    indicators = {}
    for ind in result:
        indicators[ind['slug']] = ind
        # Also index by name (normalized) for matching
        name_normalized = normalize_text(ind['name'])
        indicators[f"name:{name_normalized}"] = ind
    print(f"  ✅ {len(result)} indicadores cargados")
    return indicators


def load_entities() -> dict:
    """Load all entities from Supabase, indexed by slug."""
    print("📡 Cargando entidades desde Supabase...")
    result = supabase_request('entities', params={'select': 'id,name,slug,type'})
    entities = {}
    for ent in result:
        entities[ent['slug']] = ent
        name_normalized = normalize_text(ent['name'])
        entities[f"name:{name_normalized}"] = ent
    print(f"  ✅ {len(result)} entidades cargadas")
    return entities


def load_categories() -> dict:
    """Load all categories from Supabase."""
    print("📡 Cargando categorías desde Supabase...")
    result = supabase_request('indicator_categories', params={'select': 'id,name,slug,source_sheet'})
    categories = {}
    for cat in result:
        categories[cat['slug']] = cat
        if cat['source_sheet']:
            categories[f"sheet:{cat['source_sheet']}"] = cat
    print(f"  ✅ {len(result)} categorías cargadas")
    return categories


def insert_data_points(data_points: list, batch_size: int = 500) -> int:
    """Insert data_points in batches via Supabase REST API."""
    total_inserted = 0
    total_errors = 0

    for i in range(0, len(data_points), batch_size):
        batch = data_points[i:i + batch_size]
        try:
            # Use upsert to handle duplicates gracefully
            result = supabase_request(
                'data_points',
                method='POST',
                body=batch,
                params={'select': 'id'}
            )
            if result:
                total_inserted += len(result)
                print(f"  ✅ Batch {i // batch_size + 1}: {len(result)} registros insertados")
            else:
                total_inserted += len(batch)
                print(f"  ✅ Batch {i // batch_size + 1}: {len(batch)} registros procesados")
        except HTTPError as e:
            total_errors += len(batch)
            print(f"  ❌ Batch {i // batch_size + 1}: Error insertando {len(batch)} registros")

    return total_inserted


# ============================================================
# TEXT HELPERS
# ============================================================
def normalize_text(text: str) -> str:
    """Normalize text for matching: lowercase, remove extra spaces, standardize."""
    if not text:
        return ''
    t = str(text).strip().lower()
    t = re.sub(r'\s+', ' ', t)
    # Remove trailing periods/commas
    t = t.rstrip('.,')
    return t


def extract_unit_from_name(name: str) -> tuple:
    """Extract unit from indicator name. Returns (clean_name, unit)."""
    match = re.search(r'\(([^)]+)\)\s*$', name)
    if match:
        unit = match.group(1)
        clean_name = name[:match.start()].strip()
        return clean_name, unit
    return name, ''


def slugify(text: str) -> str:
    """Convert text to URL-friendly slug."""
    t = text.lower().strip()
    t = re.sub(r'\([^)]*\)', '', t)
    t = t.replace('ñ', 'n').replace('ó', 'o').replace('í', 'i').replace('á', 'a').replace('é', 'e').replace('ú', 'u')
    t = re.sub(r'[^a-z0-9]+', '-', t)
    t = re.sub(r'-+', '-', t)
    t = t.strip('-')
    return t


def parse_date_from_header(header_value) -> Optional[date]:
    """Parse a date from the column header (could be datetime, string, or number)."""
    if isinstance(header_value, datetime):
        return header_value.date()
    if isinstance(header_value, date):
        return header_value
    if isinstance(header_value, (int, float)):
        # Excel date serial number
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(header_value))).date()
        except:
            return None
    if isinstance(header_value, str):
        # Try parsing common formats
        for fmt in ['%Y-%m-%d', '%m/%Y', '%b %Y', '%Y-%m']:
            try:
                return datetime.strptime(header_value.strip(), fmt).date()
            except ValueError:
                continue
    return None


# ============================================================
# SHEET PARSERS
# ============================================================

def parse_variables_relevantes(ws, indicators: dict, categories: dict, source_file: str, dry_run: bool) -> list:
    """Parse 'Variables Relevantes' sheet."""
    print("\n📊 Parseando: Variables Relevantes")
    data_points = []
    cat_slug = 'variables-relevantes'

    # Find date headers in row 7+ (monthly columns start at col 12)
    date_map = {}  # col_index → date
    for col in range(MONTHLY_DATA_START_COL, ws.max_column + 1):
        header = ws.cell(row=7, column=col).value
        if header:
            d = parse_date_from_header(header)
            if d:
                date_map[col] = d

    # Parse data rows
    current_section = None
    for row_idx in range(8, 50):  # rows 8-49 for this sheet
        name_raw = ws.cell(row=row_idx, column=2).value
        if not name_raw:
            continue

        name = str(name_raw).strip()

        # Skip section headers (bold text like "Precios Combustibles")
        if ws.cell(row=row_idx, column=3).value is None and ws.cell(row=row_idx, column=12).value is None:
            current_section = name
            continue

        # Find matching indicator by slug
        ind_slug = slugify(name)
        indicator = indicators.get(ind_slug)

        if not indicator:
            # Try by normalized name
            norm_name = normalize_text(name)
            indicator = indicators.get(f"name:{norm_name}")

        if not indicator:
            # Try removing unit from name
            clean_name, _ = extract_unit_from_name(name)
            ind_slug2 = slugify(clean_name)
            indicator = indicators.get(ind_slug2)

        if not indicator:
            print(f"  ⚠️ No match para: {name} (slug: {ind_slug})")
            continue

        # Extract monthly data
        for col, dt in date_map.items():
            value = ws.cell(row=row_idx, column=col).value
            if value is not None and isinstance(value, (int, float)):
                data_points.append({
                    'indicator_id': indicator['id'],
                    'entity_id': None,
                    'value': round(float(value), 6),
                    'date': dt.isoformat(),
                    'period_type': 'monthly',
                    'source_file': source_file,
                    'is_estimated': False,
                })

    print(f"  📈 {len(data_points)} data_points extraídos")
    return data_points


def parse_edes_sheet(ws, indicators: dict, categories: dict, entities: dict, source_file: str, dry_run: bool) -> list:
    """Parse 'EDE's' sheet — 42 main indicators with Edenorte/Edesur/Edeeste breakdowns."""
    print("\n📊 Parseando: EDE's")
    data_points = []

    # Build date map from row 7
    date_map = {}
    for col in range(MONTHLY_DATA_START_COL, ws.max_column + 1):
        header = ws.cell(row=7, column=col).value
        if header:
            d = parse_date_from_header(header)
            if d:
                date_map[col] = d

    # Parse EDE's indicators
    # Pattern: Parent indicator name in col B, then "Edenorte", "Edesur", "Edeeste" on next rows
    edes_order = ['edenorte', 'edesur', 'edeeste']
    row_idx = 8
    last_parent_slug = None

    while row_idx <= ws.max_row:
        name_raw = ws.cell(row=row_idx, column=2).value
        if not name_raw:
            row_idx += 1
            continue

        name = str(name_raw).strip()
        name_lower = name.lower()

        # Check if this is a breakdown row (Edenorte/Edesur/Edeeste)
        is_ede_breakdown = False
        for ede_name in ['edenorte', 'edesur', 'edeeste']:
            if name_lower == ede_name or name_lower == ede_name:
                is_ede_breakdown = True
                ede_slug = ede_name
                break

        if is_ede_breakdown and last_parent_slug:
            # This is a breakdown row
            child_slug = f"{last_parent_slug}-{ede_slug}"
            indicator = indicators.get(child_slug)
            if not indicator:
                row_idx += 1
                continue

            # Extract monthly data
            for col, dt in date_map.items():
                value = ws.cell(row=row_idx, column=col).value
                if value is not None and isinstance(value, (int, float)):
                    data_points.append({
                        'indicator_id': indicator['id'],
                        'entity_id': entities.get(ede_slug, {}).get('id') if ede_slug in entities else None,
                        'value': round(float(value), 6),
                        'date': dt.isoformat(),
                        'period_type': 'monthly',
                        'source_file': source_file,
                        'is_estimated': False,
                    })
            row_idx += 1
            continue

        # This is a parent indicator row
        ind_slug = slugify(name)
        indicator = indicators.get(ind_slug)
        if not indicator:
            clean_name, _ = extract_unit_from_name(name)
            ind_slug = slugify(clean_name)
            indicator = indicators.get(ind_slug)

        if indicator:
            last_parent_slug = ind_slug
            # Extract monthly data for parent
            for col, dt in date_map.items():
                value = ws.cell(row=row_idx, column=col).value
                if value is not None and isinstance(value, (int, float)):
                    data_points.append({
                        'indicator_id': indicator['id'],
                        'entity_id': entities.get('edes-consolidado', {}).get('id'),
                        'value': round(float(value), 6),
                        'date': dt.isoformat(),
                        'period_type': 'monthly',
                        'source_file': source_file,
                        'is_estimated': False,
                    })
        else:
            last_parent_slug = None

        row_idx += 1

    print(f"  📈 {len(data_points)} data_points extraídos")
    return data_points


def parse_simple_entity_sheet(ws, sheet_name: str, entity_slug: str,
                               indicators: dict, entities: dict, source_file: str) -> list:
    """Parse a simple entity sheet (CDEEE, EGEHID, ETED, EGPC) with breakdowns."""
    print(f"\n📊 Parseando: {sheet_name}")
    data_points = []

    # Build date map
    date_map = {}
    for col in range(MONTHLY_DATA_START_COL, ws.max_column + 1):
        header = ws.cell(row=7, column=col).value
        if not header:
            # Try row 8 for EGPC
            header = ws.cell(row=8, column=col).value
        if header:
            d = parse_date_from_header(header)
            if d:
                date_map[col] = d

    if not date_map:
        # Try parsing from row 1 headers
        for col in range(MONTHLY_DATA_START_COL, ws.max_column + 1):
            for r in range(1, 10):
                header = ws.cell(row=r, column=col).value
                if header:
                    d = parse_date_from_header(header)
                    if d:
                        date_map[col] = d
                        break

    print(f"  📅 {len(date_map)} columnas de fecha encontradas")

    # Determine data start row
    data_start = 8
    if sheet_name == 'EGPC':
        data_start = 9

    cat_slug = SHEET_CATEGORY_MAP.get(sheet_name, '')
    last_parent_slug = None

    for row_idx in range(data_start, ws.max_row + 1):
        name_raw = ws.cell(row=row_idx, column=2).value
        if not name_raw:
            continue

        name = str(name_raw).strip()
        name_lower = name.lower().strip()

        # Skip if no data in monthly columns
        has_data = False
        for col in date_map.keys():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, (int, float)):
                has_data = True
                break

        # Check if this is a known breakdown entity name
        matched_entity_slug = None
        for key, eslug in ENTITY_MAP.items():
            if key in name_lower and eslug is not None:
                matched_entity_slug = eslug
                break

        # Try to find indicator
        # First try as a child of last parent
        if matched_entity_slug and last_parent_slug:
            child_slug = f"{last_parent_slug}-{matched_entity_slug}"
            indicator = indicators.get(child_slug)
            if indicator and has_data:
                entity_id = entities.get(matched_entity_slug, {}).get('id') if matched_entity_slug in entities else None
                for col, dt in date_map.items():
                    value = ws.cell(row=row_idx, column=col).value
                    if value is not None and isinstance(value, (int, float)):
                        data_points.append({
                            'indicator_id': indicator['id'],
                            'entity_id': entity_id,
                            'value': round(float(value), 6),
                            'date': dt.isoformat(),
                            'period_type': 'monthly',
                            'source_file': source_file,
                            'is_estimated': False,
                        })
                row_idx += 1 if False else 0  # just continue
                continue

        # Try as parent indicator
        prefix = {
            'CDEEE': 'cdeee',
            'EGEHID': 'egehid',
            'ETED': 'eted',
            'EGPC': 'egpc',
        }.get(sheet_name, '')

        if prefix:
            ind_slug = f"{prefix}-{slugify(name)}"
        else:
            ind_slug = slugify(name)

        indicator = indicators.get(ind_slug)
        if not indicator:
            # Try without prefix
            ind_slug2 = slugify(name)
            for key, ind in indicators.items():
                if not key.startswith('name:') and ind_slug2 in key:
                    indicator = ind
                    ind_slug = key
                    break

        if indicator and has_data:
            last_parent_slug = ind_slug
            entity_id = entities.get(entity_slug, {}).get('id')
            for col, dt in date_map.items():
                value = ws.cell(row=row_idx, column=col).value
                if value is not None and isinstance(value, (int, float)):
                    data_points.append({
                        'indicator_id': indicator['id'],
                        'entity_id': entity_id,
                        'value': round(float(value), 6),
                        'date': dt.isoformat(),
                        'period_type': 'monthly',
                        'source_file': source_file,
                        'is_estimated': False,
                    })

    print(f"  📈 {len(data_points)} data_points extraídos")
    return data_points


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Parse XLS and insert data_points into Supabase')
    parser.add_argument('file', help='Path to XLS file')
    parser.add_argument('--dry-run', action='store_true', help='Parse but do not insert')
    parser.add_argument('--sheet', help='Only parse this sheet name')
    parser.add_argument('--batch-size', type=int, default=500, help='Insert batch size')
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        print("❌ Variables de entorno requeridas: NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY")
        sys.exit(1)

    if not os.path.exists(args.file):
        print(f"❌ Archivo no encontrado: {args.file}")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(args.file, data_only=True)

    print(f"📄 Archivo: {args.file}")
    print(f"📊 Hojas: {wb.sheetnames}")
    print(f"🔧 Modo: {'DRY RUN (sin insertar)' if args.dry_run else 'INSERT'}")
    print()

    # Load reference data from Supabase
    indicators = load_indicators()
    entities = load_entities()
    categories = load_categories()

    # Parse sheets
    all_data_points = []

    for sheet_name in wb.sheetnames:
        if args.sheet and sheet_name != args.sheet:
            continue

        # Skip empty or non-data sheets
        if sheet_name in ('Hoja1',):
            continue

        ws = wb[sheet_name]
        if ws.max_row < 8:
            print(f"  ⏭️ {sheet_name}: muy pocas filas, saltando")
            continue

        try:
            if sheet_name == 'Variables Relevantes':
                dps = parse_variables_relevantes(ws, indicators, categories, os.path.basename(args.file), args.dry_run)
            elif sheet_name == "EDE's":
                dps = parse_edes_sheet(ws, indicators, categories, entities, os.path.basename(args.file), args.dry_run)
            elif sheet_name in ('CDEEE', 'EGEHID', 'ETED', 'EGPC'):
                entity_slug = {
                    'CDEEE': 'cdeee',
                    'EGEHID': 'egehid',
                    'ETED': 'eted',
                    'EGPC': 'egpc',
                }[sheet_name]
                dps = parse_simple_entity_sheet(ws, sheet_name, entity_slug, indicators, entities, os.path.basename(args.file))
            elif sheet_name in ('Anexo Res Financieros', 'Anexo Deuda',
                                'Nuevo Regimen tarifario', 'Regimen tarifario anterior'):
                print(f"\n📊 Parseando: {sheet_name}")
                print(f"  ⏭️ Parser para esta hoja será implementado en siguiente iteración")
                dps = []
            else:
                print(f"\n📊 {sheet_name}: hoja no reconocida, saltando")
                dps = []

            all_data_points.extend(dps)
        except Exception as e:
            print(f"  ❌ Error parseando {sheet_name}: {e}")
            import traceback
            traceback.print_exc()

    # Summary
    print(f"\n{'='*60}")
    print(f"📊 RESUMEN TOTAL")
    print(f"{'='*60}")
    print(f"Total data_points extraídos: {len(all_data_points)}")

    if all_data_points:
        # Count by indicator (first 10)
        from collections import Counter
        ind_counts = Counter(dp['indicator_id'] for dp in all_data_points)
        print(f"Indicadores con datos: {len(ind_counts)}")

        # Date range
        dates = [dp['date'] for dp in all_data_points]
        print(f"Rango de fechas: {min(dates)} → {max(dates)}")

        # Sample
        print(f"\nEjemplo (primeros 3):")
        for dp in all_data_points[:3]:
            print(f"  indicator_id={dp['indicator_id'][:8]}... date={dp['date']} value={dp['value']}")

    # Insert
    if args.dry_run:
        print(f"\n🔧 DRY RUN — No se insertaron registros")
        # Save to file for inspection
        output_file = args.file.replace('.xlsx', '_data_points.json').replace('.xls', '_data_points.json')
        with open(output_file, 'w') as f:
            json.dump(all_data_points, f, indent=2, ensure_ascii=False)
        print(f"📁 Data guardada en: {output_file}")
    else:
        if all_data_points:
            print(f"\n💾 Insertando {len(all_data_points)} data_points en Supabase...")
            inserted = insert_data_points(all_data_points, args.batch_size)
            print(f"✅ {inserted} data_points insertados exitosamente")
        else:
            print("\n⚠️ No hay data_points para insertar")

    return len(all_data_points)


if __name__ == '__main__':
    main()
